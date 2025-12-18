from tkinter import ttk
import pickle
import random
import tkinter as tk
from tkinter import *
import openpyxl
import csv
from tkinter import filedialog

#REPAIR PROGRAM--------------------------------------------------------------------------------------------------------

w = 0
s = 0
total_price = 0

#Inventory

#Check Inventory Stock
def check_inv():
    def update_stock_levels(filename):
        n = 0
        int_dis.delete("1.0", tk.END)
        try:
            wb = openpyxl.load_workbook(filename)
            sheet = wb.active

            headers = {cell.value.lower(): cell.column for cell in sheet[1]}

            if 'parts' not in headers or 'stock' not in headers:
                int_dis.insert(tk.INSERT,"Error: Could not find 'parts' or 'stock' columns.")
                return

            parts_col = headers['parts']
            stock_col = headers['stock']

            for row in range(2, sheet.max_row + 1):
                stock_cell = sheet.cell(row=row, column=stock_col)

                if stock_cell.value is not None and isinstance(stock_cell.value, (int, float)):
                    if stock_cell.value < 10:
                        stock_cell.value = 30
                        part_name = sheet.cell(row=row, column=parts_col).value
                        n += 1
                        ck = True

                    else:
                        ck = False

            if ck:
                int_dis.insert(tk.INSERT,f"Updated inventory. {n} parts restocked.")
                wb.save(filename)
                int_dis.insert(tk.INSERT, f"\nSuccessfully updated and saved {filename}")
            elif not ck:
                int_dis.insert(tk.INSERT,"Stock is up to date")

        except FileNotFoundError:
            int_dis.insert(tk.INSERT,f"Error: The file '{filename}' was not found.")
        except Exception as e:
            int_dis.insert(tk.INSERT,f"An error occurred: {e}")

    update_stock_levels('stock.xlsx')

#Parts Cost
def process_stock(filename):
    global total_price
    try:
        wb = openpyxl.load_workbook(filename)
        sheet = wb.active


            # 2. Get all data rows (excluding the header at row 1)
            # We use range starting from 2 to the max row
        max_row = sheet.max_row
        if max_row < 6:
            inv_dis.insert(tk.INSERT,"\nNot enough items in the sheet to pick 5.")
            return

            # Pick 5 unique random row indices
        random_indices = random.sample(range(2, max_row + 1), 5)

        inv_dis.insert(tk.INSERT,"\nItems selected:")

            # 3. Process the selected rows
        for row_idx in random_indices:
            part_name = sheet.cell(row=row_idx, column=1).value
            current_stock = sheet.cell(row=row_idx, column=2).value
            price = sheet.cell(row=row_idx, column=3).value

                # Update the stock (minus 1)
                # Note: In a real scenario, you might want to check if stock > 0 first
            sheet.cell(row=row_idx, column=2).value = current_stock - 1

                # Add to total
            total_price += price

            inv_dis.insert(tk.INSERT,f"\n- {part_name}: ${price} (New Stock: {current_stock - 1})")

            # 4. Save the changes back to the file
        wb.save(filename)

        inv_dis.insert(tk.INSERT, f"\nTotal Price for selection: ${total_price:.2f}")
        inv_dis.insert(tk.INSERT,f"\nChanges saved to {filename}")


    except FileNotFoundError:
        inv_dis.insert(tk.INSERT,f"\nError: The file '{filename}' was not found.")
        return

#Invoice Class

class Invoice:
    def __init__(self, inv_id = 0, name = "", dob = "", phone = "", email = "", card_name = "", card_number = "", card_expiration = "", card_ccv = "", car_make = "", car_model = "", car_year = "", car_color = "", issue = "", diag_or_repair = "", est_labor = 0):
        self.inv_id = inv_id
        self.name = name
        self.dob = dob
        self.phone = phone
        self.email = email
        self.card_name = card_name
        self.card_number = card_number
        self.card_expiration = card_expiration
        self.card_ccv = card_ccv
        self.car_make = car_make
        self.car_model = car_model
        self.car_year = car_year
        self.car_color = car_color
        self.issue = issue
        self.diag_or_repair = diag_or_repair
        self.est_labor_hrs = est_labor
        self.hrs_per_day = 8
        self.per_hour_pay = 25
        self.total_labor_cost = 0.0
        self.parts_needed = {}
        self.parts_cost = 0.0
        self.total_cost = 0.0
        self.eta_days = 0

#Create Invoice Calcs
    def invoice_calc(self):
        self.total_cost = self.labor_calc()
        self.parts_cost = self.parts_calc()
        self.total_cost = self.total_cost_calc()
        self.eta_days = self.eta_calc()

#Calculate Labor Cost
    def labor_calc(self):
        total_labor_cost = self.per_hour_pay * self.est_labor_hrs
        inv_dis.delete("1.0", tk.END)
        return total_labor_cost

#Calculate Parts Cost
    def parts_calc(self):
        parts_cost = 0
        for e in self.parts_needed:
            parts_cost += e
        return parts_cost

#Calculate Total Cost
    def total_cost_calc(self):
        total_cost = self.parts_cost + self.total_labor_cost
        return total_cost

#Calculate ETA
    def eta_calc(self):
        eta = self.est_labor_hrs / self.hrs_per_day
        return eta

#Add Invoice to Wait List
    def update_wait_list(self):
        schedule.add_schedule(self.name, self.phone, self.email)

    def display_invoice(self):
        inv_dis.insert(tk.INSERT,"\n--------------------------------------------")
        inv_dis.insert(tk.INSERT, f"\nName: {self.name} \nDOB: {self.dob} \nPhone: {self.phone} \nEmail: {self.email} \nCard Name: {self.card_name} \nCard Number: {self.card_number} \nCard Expiry: {self.card_expiration} \nCard CVV: {self.card_ccv} \nCar Make: {self.car_make} \nCar Model: {self.car_model} \nCar Color: {self.car_color} \nCar Year: {self.car_year} \nIssue: {self.issue} \nDiag or Repair: {self.diag_or_repair} \nEstimated Labor Hours: {self.est_labor_hrs} \nTotal Parts Cost: ${total_price}")
        inv_dis.insert(tk.INSERT,"\n--------------------------------------------")



#Scheduling Class

class Scheduling:
    def __init__(self):
        self.sch_id = 0
        self.name = ""
        self.phone = ""
        self.email = ""
        self.wait_list = []
        self.total_scheduled = 1

#Add to Schedule
    def add_schedule(self, name, phone, email):
        self.wait_list.append({"Name": name, "Phone": phone, "Email": email})
        sch_dis.delete("1.0", tk.END)
        sch_dis.insert(tk.INSERT,"Schedule Updated...")

#Remove Cancelled Job From Schedule
    def remove_schedule(self, x):
        for n in self.wait_list:
            if n["Name"] == x:
                self.wait_list.pop(0)

#Check Schedule
    def check_schedule(self):
        self.total_scheduled = len(self.wait_list)
        if len(self.wait_list) > 0:
            next_job = self.wait_list[0]
            sch_dis.delete("1.0", tk.END)
            sch_dis.insert(tk.INSERT, "Next Job:\nName: " + next_job["Name"] + "\nPhone: " + next_job["Phone"] + "\nEmail: " + next_job["Email"] + "\n\nTotal Scheduled: " + str(self.total_scheduled))

def create_invoice(entry1, entry2, entry3, entry4, entry5, entry6, entry7, entry8, entry9, entry10,
                           entry11, entry12, entry13, entry14, entry15, entry16):
    inv_id = entry1
    name = entry2
    dob = entry3
    phone = entry4
    email = entry5
    card_number = entry6
    card_name = entry7
    card_expiration = entry8
    card_ccv = entry9
    car_make = entry10
    car_model = entry11
    car_year = entry12
    car_color = entry13
    issue = entry14
    diag_or_repair = entry15
    est_labor_hrs = int(entry16)

    new_inv = Invoice(inv_id, name, dob, phone, email, card_name,
                              card_number, card_expiration, card_ccv, car_make, car_model,
                              car_year, car_color, issue, diag_or_repair, est_labor_hrs)
    new_inv.invoice_calc()
    invoices.append(new_inv)

    inv_dis.insert(tk.INSERT, "Invoice Created")

#SAVE FILES-------------------------------------------------------------------------------------------------------------

invent = Invoice(1,"2","3","4","5","6","7","8","9","10","11","12","13","14","15",16)
invent2 = Invoice(2, "asfd", "fwqef", "afav", "afewf", "asf", "feff", "awfw", "wv", "wewe", "weff", "giige", "regnre", "sns", "afasdf", 24)

schedule = Scheduling()
invoices = [invent, invent2]
schedules = [schedule]

#schedule.wait_list.append({"Name": "Hagan", "Phone": "5013948846", "Email": "haganzgriffin@gmail.com"})

#MAIN MENU--------------------------------------------------------------------------------------------------------------

main = Tk()
main.geometry("500x300")

top = tk.Toplevel(main)
top.geometry("700x500")
top.withdraw()

dis = Text(top, width=60, height=15)
dis.place(x=100, y=50)

#SCHEDULE WINDOW--------------------------------------------------------------------------------------------------------

sch = tk.Toplevel(top)
sch.geometry("700x500")
sch.title("Scheduling")

sch_dis = Text(sch, width=60, height=15)
sch_dis.place(x=100, y=50)

sch_ent = Entry(sch, width=60)
sch_ent.place(x=100, y=305)

sch.withdraw()

def schedule_window():

    sch.deiconify()

    def sch_show(x):
        global s
        if x == "csh":
            if len(schedule.wait_list) > 0:
                schedule.check_schedule()
            else:
                sch_dis.delete("1.0", tk.END)
                sch_dis.insert(tk.INSERT, "No jobs scheduled...")

        elif x == "addsh":
            sch_dis.delete("1.0", tk.END)
            sch_creator()

        elif x == "remsh":
            if len(schedule.wait_list) > 0:
                s = 1
                sch_dis.delete("1.0", tk.END)
                sch_dis.insert(tk.INSERT, """Enter the name to be removed from schedule\nthen click "Submit" """)
            else:
                sch_dis.delete("1.0", tk.END)
                sch_dis.insert(tk.INSERT, "No items in the schedule...")
                s = 0

        elif x == "exit":
            top.deiconify()
            sch.withdraw()

        elif x == "sub":
            get_ent = sch_ent.get()
            sch_ent.delete(0, tk.END)
            if s == 1:
                for e in schedule.wait_list:
                    if e["Name"].lower() == get_ent.lower():
                        schedule.remove_schedule(get_ent)
                        sch_dis.delete("1.0", tk.END)
                        sch_dis.insert(tk.INSERT, "Schedule item removed...")
                    else:
                        sch_dis.delete("1.0", tk.END)
                        sch_dis.insert(tk.INSERT, "No items in schedule matching search term...")
                s = 0
            elif s == 2:
                for e in schedules[0].wait_list:
                    if e["Name"] == get_ent:
                        sch_dis.delete("1.0", tk.END)
                        sch_dis.insert(tk.INSERT, "Name: " + e["Name"] + "\nPhone: " + e["Phone"] + "\nEmail: " + e["Email"])
                    else:
                        sch_dis.delete("1.0", tk.END)
                        sch_dis.insert(tk.INSERT, "No items in schedule matching search term...")
                s = 0

        elif x == "sersh":
            if len(schedule.wait_list) > 0:
                s = 2
                sch_dis.delete("1.0", tk.END)
                sch_dis.insert(tk.INSERT, """Enter the name you are searching for\nthen click "Submit" """)
            else:
                sch_dis.delete("1.0", tk.END)
                sch_dis.insert(tk.INSERT, "No items in the schedule...")
                s = 0

    ch_sh = Button(sch, text="Check Schedule", width=20, height=2, command=lambda: sch_show("csh"))
    ch_sh.place(x=100, y=330)
    add_sh = Button(sch, text="Add to Schedule", width=20, height=2, command=lambda: sch_show("addsh"))
    add_sh.place(x=100, y=380)
    rem_sh = Button(sch, text="Remove from Schedule", width=20, height=2, command=lambda: sch_show("remsh"))
    rem_sh.place(x=260, y=330)
    ex = Button(sch, text="Exit", width=15, height=2, command=lambda: sch_show("exit"))
    ex.place(x=260, y=380)
    sub = Button(sch, text="Submit", width=10, height=1, command=lambda: sch_show("sub"))
    sub.place(x=478, y=300)
    sersh = Button(sch, text="Search Schedule", width=20, height=2, command=lambda: sch_show("sersh"))
    sersh.place(x=100, y=430)

#INVOICES WINDOW--------------------------------------------------------------------------------------------------------

inv_wind = tk.Toplevel(top)
inv_wind.geometry("700x500")
inv_wind.title("Invoices")

inv_dis = Text(inv_wind, width=60, height=15)
inv_dis.place(x=100, y=50)

inv_wind.withdraw()

def invoices_window():

    inv_wind.deiconify()

    inv_ent = Entry(inv_wind, width=60)
    inv_ent.place(x=100, y=300)

    def inv_show(x):
        global w
        if x == "crin":
            inv_create_win()
            inv_wind.withdraw()
        elif x == "show":
            inv_dis.delete("1.0", tk.END)
            if len(invoices) > 0:
                inv_dis.insert(tk.INSERT, "Displaying Invoices...\n")
                for e in invoices:
                    e.display_invoice()
            else:
                inv_dis.delete("1.0", tk.END)
                inv_dis.insert(tk.INSERT, "No invoices currently saved...")
        elif x == "del":
            if len(invoices) > 0:
                w = 1
                inv_dis.delete("1.0", tk.END)
                inv_dis.insert(tk.INSERT, """Enter the name on the invoice in the entry box below then \nclick "Submit" """)
            else:
                inv_dis.delete("1.0", tk.END)
                inv_dis.insert(tk.INSERT, "No invoices currently saved...")
                w = 0

        elif x == "serin":
            if len(invoices) > 0:
                w = 2
                inv_dis.delete("1.0", tk.END)
                inv_dis.insert(tk.INSERT, """Enter the name on the invoice in the entry box below then \nclick "Search Invoices" """)
            else:
                inv_dis.delete("1.0", tk.END)
                inv_dis.insert(tk.INSERT, "No invoices currently saved...")
                w = 0

        elif x == "submit":
            nm = inv_ent.get()
            inv_ent.delete(0, tk.END)
            if w == 1:
                for e in invoices:
                    if e.name.lower() == nm.lower():
                        invoices.remove(e)
                        inv_dis.delete("1.0", tk.END)
                        inv_dis.insert(tk.INSERT, f"""Invoice matching name "{nm}" deleted...""")
                    elif e.name.lower() != nm.lower():
                        inv_dis.delete("1.0", tk.END)
                        inv_dis.insert(tk.INSERT, f"""No invoice found matching name "{nm}"... """)
                w = 0
            elif w == 2:
                for e in invoices:
                    if e.name.lower() == nm.lower():
                        e.display_invoice()
                    elif e.name.lower() != nm.lower():
                        inv_dis.delete("1.0", tk.END)
                        inv_dis.insert(tk.INSERT, "No invoices matching that name...")
                w = 0
            else:
                inv_dis.delete("1.0", tk.END)
                inv_dis.insert(tk.INSERT, "Error")

        elif x == "exit":
            top.deiconify()
            inv_wind.withdraw()

    cr_in = Button(inv_wind, text="Create Invoice", width=20, height=2, command=lambda: inv_show("crin"))
    cr_in.place(x=100, y=325)

    sh_in = Button(inv_wind, text="Show Invoices", width=20, height=2, command=lambda: inv_show("show"))
    sh_in.place(x=100, y=375)

    del_in = Button(inv_wind, text="Delete Invoice", width=20, height=2, command=lambda: inv_show("del"))
    del_in.place(x=260, y=325)

    exit_in = Button(inv_wind, text="Exit", width=15, height=2, command=lambda: inv_show("exit"))
    exit_in.place(x=260, y=375)

    submit = Button(inv_wind, text="Submit", width=12, height=1, command=lambda: inv_show("submit"))
    submit.place(x=478, y=298)

    ser_in = Button(inv_wind, text="Search Invoices", width=20, height=2, command=lambda: inv_show("serin"))
    ser_in.place(x=100, y=425)

#INVENTORY WINDOW-------------------------------------------------------------------------------------------------------

int_men = tk.Toplevel(top)
int_men.geometry("700x500")
int_men.title("Inventory")

int_dis = Text(int_men, width=60, height=15)
int_dis.place(x=100, y=50)

int_men.withdraw()

def inventory_menu():

    int_men.deiconify()

    def int_show(x):
        if x == "intcheck":
            check_inv()

        elif x == "exit":
            int_men.withdraw()
            top.deiconify()

    int_check = Button(int_men, text="Check Inventory", width=20, height=2, command=lambda: int_show("intcheck"))
    int_check.place(x=100, y=305)

    int_exit = Button(int_men, text="Exit", width=15, height=2, command=lambda: int_show("exit"))
    int_exit.place(x=100, y=355)

#INVOICE CREATOR WINDOW-------------------------------------------------------------------------------------------------

top1 = tk.Toplevel(top)
top1.title("Invoice Creation")
top1.geometry("500x575")

top1.withdraw()

def inv_create_win():

    top1.deiconify()

    def cr_i():

        try:
            create_invoice(inv_id_entry.get(), name_entry.get(), dob_entry.get(), phone_entry.get(), email_entry.get(), card_num_entry.get(),
                               card_name_entry.get(), card_exp_entry.get(), card_cvv_entry.get(), car_make_entry.get(), car_model_entry.get(),
                               car_year_entry.get(), car_color_entry.get(), issue_entry.get(), diag_repair_entry.get(), labor_hours_entry.get())

            process_stock("stock.xlsx")

        except ValueError:
            inv_dis.delete("1.0", tk.END)
            inv_dis.insert(tk.INSERT, "One or more entry boxes were left blank...\nPlease try again...")


        top1.withdraw()
        inv_wind.deiconify()

    tk.Label(top1, text="Invoice ID:").place(x=10, y=10)
    inv_id_entry = Entry(top1, width=60)
    inv_id_entry.place(x=125, y=12)

    tk.Label(top1, text="Name:").place(x=10, y=40)
    name_entry = Entry(top1, width=60)
    name_entry.place(x=125, y=42)

    tk.Label(top1, text="DOB:").place(x=10, y=70)
    dob_entry = Entry(top1, width=60)
    dob_entry.place(x=125, y=72)

    tk.Label(top1, text="Phone:").place(x=10, y=100)
    phone_entry = Entry(top1, width=60)
    phone_entry.place(x=125, y=102)

    tk.Label(top1, text="Email:").place(x=10, y=130)
    email_entry = Entry(top1, width=60)
    email_entry.place(x=125, y=132)

    tk.Label(top1, text="Card Number:").place(x=10, y=160)
    card_num_entry = Entry(top1, width=60)
    card_num_entry.place(x=125, y=162)

    tk.Label(top1, text="Card Name:").place(x=10, y=190)
    card_name_entry = Entry(top1, width=60)
    card_name_entry.place(x=125, y=192)

    tk.Label(top1, text="Card Expiration:").place(x=10, y=220)
    card_exp_entry = Entry(top1, width=60)
    card_exp_entry.place(x=125, y=222)

    tk.Label(top1, text="Card CVV:").place(x=10, y=250)
    card_cvv_entry = Entry(top1, width=60)
    card_cvv_entry.place(x=125, y=252)

    tk.Label(top1, text="Car Make:").place(x=10, y=280)
    car_make_entry = Entry(top1, width=60)
    car_make_entry.place(x=125, y=282)

    tk.Label(top1, text="Car Model:").place(x=10, y=310)
    car_model_entry = Entry(top1, width=60)
    car_model_entry.place(x=125, y=312)

    tk.Label(top1, text="Car Year:").place(x=10, y=340)
    car_year_entry = Entry(top1, width=60)
    car_year_entry.place(x=125, y=342)

    tk.Label(top1, text="Car Color:").place(x=10, y=370)
    car_color_entry = Entry(top1, width=60)
    car_color_entry.place(x=125, y=372)

    tk.Label(top1, text="Issue:").place(x=10, y=400)
    issue_entry = Entry(top1, width=60)
    issue_entry.place(x=125, y=402)

    tk.Label(top1, text="Diagnosis or Repair:").place(x=10, y=430)
    diag_repair_entry = Entry(top1, width=60)
    diag_repair_entry.place(x=125, y=432)

    tk.Label(top1, text="Est. Labor Hours:").place(x=10, y=460)
    labor_hours_entry = Entry(top1, width=60)
    labor_hours_entry.place(x=125, y=462)

    submit = Button(top1, text="Submit", width=20, height=2, command=cr_i)
    submit.place(x=175, y=500)

#SCHEDULE CREATOR-------------------------------------------------------------------------------------------------------

sh_cr = tk.Toplevel(top)
sh_cr.geometry("500x150")
sh_cr.title("Schedule Creation")

sh_cr.withdraw()

def sch_creator():

    sh_cr.deiconify()

    def sh_show(x):
        if x == "sub":
            if name_entry.get() == "" or phone_entry.get() == "" or email_entry.get() == "":
                sch_dis.delete("1.0", tk.END)
                sch_dis.insert(tk.INSERT, "One or more boxes were left blank...\nPlease try again...")
            else:
                schedule.add_schedule(name_entry.get(), phone_entry.get(), email_entry.get())

            sh_cr.withdraw()

    tk.Label(sh_cr, text="Name").place(x=10, y=10)
    name_entry = Entry(sh_cr, width=60)
    name_entry.place(x=125, y=12)

    tk.Label(sh_cr, text="Phone").place(x=10, y=40)
    phone_entry = Entry(sh_cr, width=60)
    phone_entry.place(x=125, y=42)

    tk.Label(sh_cr, text="Email").place(x=10, y=70)
    email_entry = Entry(sh_cr, width=60)
    email_entry.place(x=125, y=72)

    submit = Button(sh_cr, text="Submit", width=20, height=1, command=lambda: sh_show("sub"))
    submit.place(x=175, y=102)

#MAIN WINDOW BUTTONS----------------------------------------------------------------------------------------------------

def show(x):
        if x == "ci":
            invoices_window()
            top.withdraw()

        elif x == "cin":
            inventory_menu()
            top.withdraw()

        elif x == "sch":
            schedule_window()
            top.withdraw()

        elif x == "save":
            dis.delete("1.0", tk.END)
            dis.insert(tk.INSERT, "Saving...\nSaved...")

            with open("saved_invoices.pk1", "wb") as i:
                for e in invoices:
                    # noinspection PyTypeChecker
                    pickle.dump(e, i, pickle.HIGHEST_PROTOCOL)

            with open("saved_schedule.pk1", "wb") as ss:
                # noinspection PyTypeChecker
                pickle.dump(schedules[0], s, pickle.HIGHEST_PROTOCOL)

            with open("invoice_num.dat", "wb") as num:
                # noinspection PyTypeChecker
                pickle.dump(len(invoices), num)

            i.close()
            ss.close()
            num.close()

        elif x == "load":
            dis.delete("1.0", tk.END)
            dis.insert(tk.INSERT, "Loading...\nLoaded...")

            invoices.clear()
            schedules.clear()
            with open("invoice_num.dat", "rb") as num:
                inv_num = pickle.load(num)

            while True:
                try:

                    with open("saved_invoices.pk1", "rb") as i:
                        loaded_invoices = pickle.load(i)
                        invoices.append(loaded_invoices)
                        while inv_num > 0:
                            invoices.append(pickle.load(i))
                            inv_num -= 1



                    with open("saved_schedule.pk1", "rb") as ss:
                         schedules.append(pickle.load(ss))

                except EOFError:
                    break

        elif x == "exit":
            top.withdraw()
            main.deiconify()
        else:
            dis.insert(tk.INSERT, "\n\nInvalid Input")

inv_men = Button(top, text="Invoice Menu", width=20, height=2, command=lambda: show("ci"))
inv_men.place(x=110, y=305)
ch_in = Button(top, text="Inventory Menu", width=20, height=2, command=lambda: show("cin"))
ch_in.place(x=110, y=355)
sh = Button(top, text="Scheduling Menu", width=20, height=2, command=lambda: show("sch"))
sh.place(x=110, y=405)

save = Button(top, text="Save", width=15, height=2, command=lambda: show("save"))
save.place(x=270, y=305)
load = Button(top, text="Load", width=15, height=2, command=lambda: show("load"))
load.place(x=270, y=355)
exi = Button(top, text="Exit", width=15, height=2, command=lambda: show("exit"))
exi.place(x=270, y=405)


#DEALERSHIP MENU--------------------------------------------------------------------------------------------------------

totalSalesFinalVariableIDEK = 0  # why did i put this at the top idecatp


class Car:
    def __init__(self, make, model, year, color, price, repBool, count=1):
        self.make = make
        self.model = model
        self.year = year
        self.color = color
        self.price = price
        self.repBool = repBool
        self.count = count

    def __str__(self):
        return f'{self.year} {self.make} {self.model} ({self.color}) | ${self.price:,.2f} | Requires Repairs: {self.repBool}'


class Inventory:
    def __init__(self):
        self.cars = []
        self.sales = []

    def addCar(self, make, model, year, color, price, repBool, count):

        newCar = Car(make, model, year, color, price, repBool, count)
        self.cars.append(newCar)
        print(f'Vehicle Added: {color} {year} {make} {model} | ${price:,.2f} | Requires Repairs: {repBool}')

    def displayInv(self):
        print('\nCurrent Inventory:')
        if not self.cars:
            print('No cars in inventory.')
        else:
            for i, car in enumerate(self.cars, start=1):
                print(f'({i}) {car}')
        print()

    def sell(self, index):
        if 0 <= index < len(self.cars):
            car = self.cars[index]
            if car.count >= 1:
                totalPrice = car.price
                print(f'Sold {car.make} {car.model}(s) for ${totalPrice:,.2f}')
                self.sales.append(f"{car.year} {car.make} {car.model} ({car.color}) for ${car.price:,.2f}")
                self.cars.remove(car)
                return totalPrice
            else:
                print('Invalid number of vehicles in stock.')
                return 0
        else:
            print('Invalid car selection.')
            return 0

    def sellCar(self, make, model, year, color, price, repBoolYN):
        tempBool = (repBoolYN == 'Y')
        for car in self.cars:
            if (car.make == make and car.model == model and car.year == year and
                    car.color == color and car.price == price and car.repBool == tempBool):
                self.cars.remove(car)
                self.sales.append(car)
                print(f'Sold {car}')
                return car
        return None


class DealershipOOP:
    def __init__(self):
        self.inventory = Inventory()
        self.totalSales = 0


app = DealershipOOP()



root = tk.Toplevel(main)
root.title('Dealership Application')
root.geometry('800x600')
root.resizable(False, False)
root.withdraw()

notebook = ttk.Notebook(root)
notebook.pack(expand=True, fill='both')

# Frame Creation:
AddFrame = tk.Frame(notebook, width=500, height=500)
InventoryFrame = tk.Frame(notebook, width=500, height=500)

SalesFrame = tk.Frame(notebook, width=500, height=500)

AddFrame.pack()
InventoryFrame.pack()

SalesFrame.pack()

notebook.add(AddFrame, text='Add Car')
notebook.add(InventoryFrame, text='Inventory')

notebook.add(SalesFrame, text='Sales History')

def saveSales():
    if not app.inventory.sales:
        tk.messagebox.showinfo('Info', 'No sales to save.')
        return
    filePath = filedialog.asksaveasfilename(defaultextension='.csv', filetypes=[('CSV Files', '*.csv')], title='Save Sales As CSV')

    if not filePath:
        return

    with open(filePath, 'w', newline='') as file:
        writer = csv.writer(file)
        writer.writerow(["Make", "Model", "Year", "Color", "Price", "Needs Repairs"])
        for car in app.inventory.sales:
            rep = 'Y' if car.repBool else 'N'
            writer.writerow([car.make, car.model, car.year, car.color, f'{car.price:.2f}', rep])
        writer.writerow([f'Total Sales: ${totalSalesFinalVariableIDEK:,.2f}'])

    tk.messagebox.showinfo('Success', f'Saved Sales History to {filePath}.')

def sellOnSelect():
    selected_item = tree.selection()
    if selected_item:
        item_id = selected_item[0]
        values = tree.item(item_id, "values")

        make = values[0]
        model = values[1]
        year = values[2]
        color = values[3]
        price = float(values[4].replace('$', '').replace(',', ''))
        repBoolYN = values[5]

        app.inventory.sellCar(make, model, year, color, price, repBoolYN)
        updateTab(None)

def populateInventoryTree(cars):
    for item in tree.get_children():
        tree.delete(item)
    for car in cars:
        check = 'Y' if car.repBool else 'N'
        tree.insert(
            '', tk.END, text=car,
            values=(car.make, car.model, car.year, car.color, f'${car.price:,.2f}', check)
        )

def updateTab(event):
    selectedTab = notebook.select()
    print(f'Selected tab: {selectedTab}')
    global totalSalesFinalVariableIDEK
    if selectedTab == '.!notebook.!frame2':
        populateInventoryTree(app.inventory.cars)
    elif selectedTab == '.!notebook.!frame3':
        for widget in Salestree.winfo_children():
            widget.destroy()
        for item in Salestree.get_children():
            Salestree.delete(item)
        for car in app.inventory.sales:
            if car.repBool:
                check = 'Y'
            else:
                check = 'N'
            Salestree.insert('', tk.END, text=car,
                             values=(car.make, car.model, car.year, car.color, f'${car.price:,.2f}', check))
            totalSalesFinalVariableIDEK += car.price
            totalSalesLabel.config(text=f'Total Sales: ${totalSalesFinalVariableIDEK:,.2f}')

notebook.bind('<<NotebookTabChanged>>', updateTab)

# <--------------------Add Car Frame------------------>:
makeVar = tk.StringVar()
modelVar = tk.StringVar()
yearVar = tk.StringVar()
colorVar = tk.StringVar()
priceVar = tk.StringVar()
repairCheck = tk.IntVar()

def submit():
    if repairCheck.get() == 1:
        repBool = True
    else:
        repBool = False
    make = makeVar.get()
    model = modelVar.get()
    year = yearVar.get()
    color = colorVar.get()
    price = float(priceVar.get())
    app.inventory.addCar(make, model, year, color, price, repBool, 1)
    makeEntry.delete(0, tk.END)
    modelEntry.delete(0, tk.END)
    yearEntry.delete(0, tk.END)
    colorEntry.delete(0, tk.END)
    priceEntry.delete(0, tk.END)
    repairCheckbox.deselect()

def loadDefaults():
    dummyCars = [
        ("Toyota", "Camry", "2019", "Silver", 18500, False),
        ("Honda", "Civic", "2020", "Blue", 19500, False),
        ("Ford", "F-150", "2018", "Red", 27000, True),
        ("Chevrolet", "Malibu", "2017", "Black", 15000, False),
        ("Tesla", "Model 3", "2021", "White", 38000, False),
        ("BMW", "X5", "2019", "Gray", 45000, True),
        ("Audi", "A4", "2018", "Black", 28000, False),
        ("Nissan", "Altima", "2020", "Blue", 21000, False),
        ("Hyundai", "Tucson", "2017", "Red", 17000, True),
        ("Kia", "Sorento", "2019", "Silver", 25000, False),
        ("Subaru", "Outback", "2018", "Green", 23000, False),
        ("Volkswagen", "Jetta", "2016", "White", 12000, True),
        ("Dodge", "Charger", "2020", "Black", 30000, False),
        ("Jeep", "Wrangler", "2019", "Yellow", 34000, True),
        ("Mazda", "CX-5", "2021", "Red", 29000, False)
    ]

    for make, model, year, color, price, repBool in dummyCars:
        app.inventory.addCar(make, model, year, color, price, repBool, 1)

def ex():
    root.withdraw()
    main.deiconify()

makeLabel = tk.Label(AddFrame, text='Make: ', font=('calibre', 10, 'bold'))
makeEntry = tk.Entry(AddFrame, textvariable=makeVar, font=('calibre', 10, 'normal'))

modelLabel = tk.Label(AddFrame, text='Model: ', font=('calibre', 10, 'bold'))
modelEntry = tk.Entry(AddFrame, textvariable=modelVar, font=('calibre', 10, 'normal'))

yearLabel = tk.Label(AddFrame, text='Year: ', font=('calibre', 10, 'bold'))
yearEntry = tk.Entry(AddFrame, textvariable=yearVar, font=('calibre', 10, 'normal'))

colorLabel = tk.Label(AddFrame, text='Color: ', font=('calibre', 10, 'bold'))
colorEntry = tk.Entry(AddFrame, textvariable=colorVar, font=('calibre', 10, 'normal'))

priceLabel = tk.Label(AddFrame, text='Price: $', font=('calibre', 10, 'bold'))
priceEntry = tk.Entry(AddFrame, textvariable=priceVar, font=('calibre', 10, 'normal'))

repairCheckbox = tk.Checkbutton(AddFrame, text='Requires Repairs', variable=repairCheck)

submitButton = tk.Button(AddFrame, text='Submit', command=submit)
loadDefaultsButton = tk.Button(AddFrame, text='Load Defaults: Debugging Only',
                               bg='lightgray', command=loadDefaults)
loadDefaultsButton.grid(row=8, column=5, pady=10)
exit_button = tk.Button(AddFrame, text='Exit', command=ex)
exit_button.grid(row=9, column=5, pady=10)

makeLabel.grid(row=1, column=4)
makeEntry.grid(row=1, column=5)

modelLabel.grid(row=2, column=4)
modelEntry.grid(row=2, column=5)

yearLabel.grid(row=3, column=4)
yearEntry.grid(row=3, column=5)

colorLabel.grid(row=4, column=4)
colorEntry.grid(row=4, column=5)

priceLabel.grid(row=5, column=4)
priceEntry.grid(row=5, column=5)
repairCheckbox.grid(row=6, column=5)

submitButton.grid(row=7, column=5)

# <--------------------Inventory Frame------------------>:

SearchFrame = tk.LabelFrame(InventoryFrame, text="Search / Filter Cars")
SearchFrame.pack(fill="x", padx=8, pady=5)

makeSearch = tk.StringVar()
modelSearch = tk.StringVar()
yearSearch = tk.StringVar()
colorSearch = tk.StringVar()
repairSearch = tk.StringVar(value="Any")

tk.Label(SearchFrame, text="Make").grid(row=0, column=0)
tk.Entry(SearchFrame, textvariable=makeSearch, width=10).grid(row=0, column=1)

tk.Label(SearchFrame, text="Model").grid(row=0, column=2)
tk.Entry(SearchFrame, textvariable=modelSearch, width=10).grid(row=0, column=3)

tk.Label(SearchFrame, text="Year").grid(row=0, column=4)
tk.Entry(SearchFrame, textvariable=yearSearch, width=8).grid(row=0, column=5)

tk.Label(SearchFrame, text="Color").grid(row=0, column=6)
tk.Entry(SearchFrame, textvariable=colorSearch, width=10).grid(row=0, column=7)

tk.Label(SearchFrame, text="Needs Repairs").grid(row=0, column=8)
ttk.Combobox(
    SearchFrame,
    textvariable=repairSearch,
    values=["Any", "Yes", "No"],
    width=8,
    state="readonly"
).grid(row=0, column=9)

columns = ("Make", "Model", "Year", 'Color', 'Price', 'Needs Repairs')
tree = ttk.Treeview(InventoryFrame, columns=columns, show="headings", selectmode="browse")
for col in columns:
    tree.heading(col, text=col)
    if col == "Make":
        tree.column(col, width=100, anchor="center")
    elif col == "Model":
        tree.column(col, width=100, anchor="center")
    elif col == 'Needs Repairs':
        tree.column(col, width=50, anchor='center')
    else:
        tree.column(col, width=100, anchor="center")
tree.pack(fill="both", expand=True, padx=8, pady=6)

sellButton = tk.Button(InventoryFrame, text='SELL', bg='red', height=2, width=10, command=sellOnSelect)
sellButton.pack(padx=10, pady=10, side=tk.LEFT, anchor='w')

def applyFilters():
    filteredCars = []
    for car in app.inventory.cars:
        if makeSearch.get() and makeSearch.get().lower() not in car.make.lower(): continue
        if modelSearch.get() and modelSearch.get().lower() not in car.model.lower(): continue
        if yearSearch.get() and yearSearch.get() != str(car.year): continue
        if colorSearch.get() and colorSearch.get().lower() not in car.color.lower(): continue
        if repairSearch.get() == "Yes" and not car.repBool: continue
        if repairSearch.get() == "No" and car.repBool: continue
        filteredCars.append(car)
    populateInventoryTree(filteredCars)

def resetFilters():
    makeSearch.set("")
    modelSearch.set("")
    yearSearch.set("")
    colorSearch.set("")
    repairSearch.set("Any")
    populateInventoryTree(app.inventory.cars)

tk.Button(SearchFrame, text="Search", command=applyFilters).grid(row=0, column=10, padx=5)
tk.Button(SearchFrame, text="Reset", command=resetFilters).grid(row=0, column=11)

# <--------------------Sales Frame------------------>:
Salestree = ttk.Treeview(SalesFrame, columns=columns, show="headings", selectmode="browse")
for col in columns:
    Salestree.heading(col, text=col)
    if col == "Make":
        Salestree.column(col, width=100, anchor="center")
    elif col == "Model":
        Salestree.column(col, width=100, anchor="center")
    elif col == 'Needs Repairs':
        Salestree.column(col, width=50, anchor='center')
    else:
        Salestree.column(col, width=100, anchor="center")
Salestree.pack(fill="both", expand=True, padx=8, pady=6)

totalSalesLabel = tk.Label(SalesFrame, text=f'Total Sales: ${totalSalesFinalVariableIDEK:,.2f}')
totalSalesLabel.pack(padx=10, pady=10, anchor='center')

saveCSVButton = tk.Button(SalesFrame, text="Save Sales as CSV", command=saveSales)
saveCSVButton.pack(pady=10)


#MAIN MENU--------------------------------------------------------------------------------------------------------------

def main_show(x):
    if x == "del":
        main.withdraw()
        root.deiconify()
    elif x == "rep":
        main.withdraw()
        top.deiconify()
    elif x == "ex":
        main.quit()

dealer = Button(main, text="Dealership", width=20, height=8, background="Grey", command=lambda: main_show("del"))
dealer.place(x=50, y=100)
repair = Button(main, text="Repair Shop", width=20, height=8, background="Gray", command=lambda: main_show("rep"))
repair.place(x=295, y=100)
main_exit = Button(main, text="Exit", width=15, height=2, background="red", command=lambda: main_show("ex"))
main_exit.place(x=190, y=250)

main.mainloop()